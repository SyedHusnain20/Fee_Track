"""Attendance archive export — Step 11 (Section 4 of the spec). Builds the
in-memory workbook (one sheet per class + a Teachers sheet) and computes
the preview summary shown before an admin confirms the archive run.
Nothing here touches the database beyond reading — deletion only happens
in app/api/archive.py, and only after a confirmed B2 upload.

Every read in this module takes an explicit max_id cutoff (the highest
AttendanceRecord.id that existed when the archive run started, captured
once by app/api/archive.py and threaded through get_archive_summary() and
build_workbook() alike) rather than reading "whatever's in the table right
now." The kiosk is unauthenticated and always available, and a B2 upload
is a real network round trip -- without a fixed cutoff, a scan landing
between the workbook being built and the final DELETE running would be
in neither: not in the exported file (already serialized before that scan
happened) and not in the live table (an unconditional DELETE has no way
to know it arrived too late to be included). Filtering every step to
`id <= max_id` means a late-arriving scan simply isn't touched at all —
it's exported (and cleared) on the *next* archive run instead, which is
correct, rather than silently vanishing on this one.
"""

from io import BytesIO

from openpyxl import Workbook
from sqlalchemy import func
from sqlmodel import Session, select

from app.models.attendance_record import AttendanceRecord
from app.models.class_level import ClassLevel
from app.models.enums import AttendanceSession
from app.models.student import Student
from app.models.teacher import Teacher

SESSION_LABELS = {
    AttendanceSession.SCHOOL: "School",
    AttendanceSession.ACADEMY: "Academy",
}

HEADER_STUDENT = ["Roll #", "Name", "Date", "Session", "Arrival", "Status"]
HEADER_TEACHER = ["Staff ID", "Name", "Date", "Session", "Arrival", "Status"]


def _status_label(record: AttendanceRecord) -> str:
    if record.punctuality_status is None:
        return "Logged"
    return "Late" if record.punctuality_status.value == "late" else "On time"


def get_max_attendance_id(session: Session) -> int:
    """The cutoff for an archive run — call this ONCE at the very start
    (before anything else touches AttendanceRecord) and thread the same
    value through get_archive_summary(), build_workbook(), and the final
    DELETE in app/api/archive.py. See module docstring for why a fixed
    cutoff matters here."""
    return session.exec(select(func.max(AttendanceRecord.id))).one() or 0


def get_archive_summary(session: Session, max_id: int) -> dict:
    """Preview data shown before an admin confirms: total record count,
    the date range covered, per-class breakdown, and teacher count —
    scoped to id <= max_id, see module docstring."""
    total = session.exec(
        select(func.count()).select_from(AttendanceRecord).where(AttendanceRecord.id <= max_id)
    ).one()
    if not total:
        return {
            "total": 0,
            "start_date": None,
            "end_date": None,
            "class_counts": [],
            "teacher_count": 0,
        }

    start_date, end_date = session.exec(
        select(func.min(AttendanceRecord.scan_date), func.max(AttendanceRecord.scan_date)).where(
            AttendanceRecord.id <= max_id
        )
    ).first()

    class_levels = session.exec(select(ClassLevel).order_by(ClassLevel.class_offset)).all()
    class_counts = []
    for cl in class_levels:
        count = session.exec(
            select(func.count())
            .select_from(AttendanceRecord)
            .join(Student, AttendanceRecord.student_id == Student.id)
            .where(Student.class_level_id == cl.id, AttendanceRecord.id <= max_id)
        ).one()
        if count:
            class_counts.append({"class_level": cl, "count": count})

    teacher_count = session.exec(
        select(func.count())
        .select_from(AttendanceRecord)
        .where(AttendanceRecord.teacher_id.is_not(None), AttendanceRecord.id <= max_id)
    ).one()

    return {
        "total": total,
        "start_date": start_date,
        "end_date": end_date,
        "class_counts": class_counts,
        "teacher_count": teacher_count,
    }


def build_workbook(session: Session, max_id: int) -> BytesIO:
    """One sheet per ClassLevel (created even if empty for that class, per
    spec's 'one sheet per class' — 15 sheets always present), plus a
    Teachers sheet. Returns an in-memory .xlsx — never written to disk
    here, so this function alone has no cleanup burden. Scoped to
    id <= max_id — see module docstring."""
    wb = Workbook()
    wb.remove(wb.active)  # drop the default blank sheet

    class_levels = session.exec(select(ClassLevel).order_by(ClassLevel.class_offset)).all()
    for cl in class_levels:
        ws = wb.create_sheet(title=cl.name[:31])  # Excel's 31-char sheet name limit
        ws.append(HEADER_STUDENT)
        rows = session.exec(
            select(AttendanceRecord, Student)
            .join(Student, AttendanceRecord.student_id == Student.id)
            .where(Student.class_level_id == cl.id, AttendanceRecord.id <= max_id)
            .order_by(AttendanceRecord.scan_date, AttendanceRecord.arrival_time)
        ).all()
        for record, student in rows:
            ws.append(
                [
                    student.roll_number,
                    student.name,
                    record.scan_date.isoformat(),
                    SESSION_LABELS[record.session],
                    record.arrival_time.strftime("%H:%M:%S"),
                    _status_label(record),
                ]
            )

    ws = wb.create_sheet(title="Teachers")
    ws.append(HEADER_TEACHER)
    teacher_rows = session.exec(
        select(AttendanceRecord, Teacher)
        .join(Teacher, AttendanceRecord.teacher_id == Teacher.id)
        .where(AttendanceRecord.id <= max_id)
        .order_by(AttendanceRecord.scan_date, AttendanceRecord.arrival_time)
    ).all()
    for record, teacher in teacher_rows:
        ws.append(
            [
                teacher.staff_id,
                teacher.name,
                record.scan_date.isoformat(),
                SESSION_LABELS[record.session],
                record.arrival_time.strftime("%H:%M:%S"),
                _status_label(record),
            ]
        )

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
